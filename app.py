from flask import Flask, render_template_string, request, send_file, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import stripe, ezdxf, fitz, os, pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config.update(SECRET_KEY='key999', SQLALCHEMY_DATABASE_URI='sqlite:///users.db')
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

stripe.api_key = "sk_test_YOUR_STRIPE_SECRET_KEY"
STRIPE_PRICE_ID = "price_YOUR_STRIPE_PLAN_ID" 
YOUR_DOMAIN = "https://YOUR_APP_://onrender.com" 

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    is_premium = db.Column(db.Boolean, default=False)

@login_manager.user_loader
def load_user(uid): return User.query.get(int(uid))

STYLE = "<style>body{font-family:'Segoe UI';background:#f4f6f9;text-align:center;padding:50px;}.container{background:white;max-width:500px;margin:auto;padding:40px;border-radius:15px;box-shadow:0 4px 15px rgba(0,0,0,0.1);}input{width:100%;padding:12px;margin:10px 0;border:1px solid #ccc;border-radius:5px;text-align:right;}.btn{background:#3498db;color:white;border:none;padding:12px;font-weight:bold;border-radius:5px;cursor:pointer;width:100%;margin-top:10px;}.btn:hover{background:#2980b9;}.btn-p{background:#f1c40f;color:#2c3e50;}.btn-d{background:#e74c3c;padding:5px 15px;font-size:12px;color:white;text-decoration:none;border-radius:3px;}.flash{color:#e74c3c;margin-bottom:15px;}.btn-u{border:2px dashed #3498db;background:#f8fafc;padding:40px 20px;border-radius:10px;font-weight:bold;cursor:pointer;display:block;width:100%;}@keyframes spin{0%{transform:rotate(0deg);}100%{transform:rotate(360deg);}}</style>"

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        uname = request.form.get('username')
        pwd = request.form.get('password')
        if User.query.filter_by(username=uname).first():
            flash('اسم المستخدم مسجل مسبقاً!')
            return redirect(url_for('register'))
        db.session.add(User(username=uname, password=generate_password_hash(pwd, method='pbkdf2:sha256')))
        db.session.commit()
        return redirect(url_for('login'))
    return render_template_string(f"<html lang='ar' dir='rtl'><head>{STYLE}</head><body><div class='container'><h1>إنشاء حساب جديد</h1>{{% with m = get_flashed_messages() %}}{{% if m %}}<div class='flash'>{{{{m[0]}}}}</div>{% endif %}{{% endwith %}}<form method='POST'><input type='text' name='username' placeholder='اسم المستخدم' required><input type='password' name='password' placeholder='كلمة المرور' required><button type='submit' class='btn'>إنشاء الحساب</button></form><br><a href='{url_for('login')}'>لديك حساب؟ سجل دخولك</a></div></body></html>")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        uname = request.form.get('username')
        pwd = request.form.get('password')
        user = User.query.filter_by(username=uname).first()
        if user and check_password_hash(user.password, pwd):
            login_user(user)
            return redirect(url_for('index'))
        flash('اسم المستخدم أو كلمة المرور غير صحيحة!')
    return render_template_string(f"<html lang='ar' dir='rtl'><head>{STYLE}</head><body><div class='container'><h1>تسجيل الدخول للمنصة</h1>{{% with m = get_flashed_messages() %}}{{% if m %}}<div class='flash'>{{{{m[0]}}}}</div>{% endif %}{{% endwith %}}<form method='POST'><input type='text' name='username' placeholder='اسم المستخدم' required><input type='password' name='password' placeholder='كلمة المرور' required><button type='submit' class='btn'>دخول</button></form><br><a href='{url_for('register')}'>ليس لديك حساب؟ اشترك الآن</a></div></body></html>")

@app.route('/logout')
@login_required
def logout(): logout_user(); return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    if not current_user.is_premium:
        return render_template_string(f"<html lang='ar' dir='rtl'><head>{STYLE}</head><body><div style='text-align:left;max-width:500px;margin:auto;'><a href='{url_for('logout')}' class='btn-d'>خروج</a></div><div class='container'><h1>حسابك الحالي مجاني ومحدود 🛑</h1><p>للوصول الكامل واستخراج جداول الكميات غير المحدودة لملفات الأوتوكاد والـ PDF، يرجى الاشتراك.</p><h2>29$ / شهرياً</h2><form action='/create-checkout-session' method='POST'><button type='submit' class='btn btn-p'>💳 اشترك الآن عبر الفيزا</button></form></div></body></html>")
    return render_template_string(f"<html lang='ar' dir='rtl'><head>{STYLE}</head><body><div style='text-align:left;max-width:550px;margin:auto;'>مرحباً، <b>{current_user.username}</b> ⭐ باقة مميزة | <a href='{url_for('logout')}' class='btn-d'>خروج</a></div><div class='container'><h1>منصة الحصر الذكي الاحترافية</h1><p>ارفع ملف المشروع (DXF أو PDF) وسيقوم الموقع بالحصر وتحميل ملف الـ Excel فوراً.</p><form id='f' action='/upload' method='post' enctype='multipart/form-data'><button type='button' class='btn-u' onclick='document.getElementById(\"i\").click()'>📁 اضغط هنا لاختيار ملف المخطط</button><input type='file' id='i' name='file' accept='.dxf,.pdf' style='display:none' onchange='document.getElementById(\"f\").submit()'></form></div></body></html>")

@app.route('/create-checkout-session', methods=['POST'])
@login_required
def create_checkout_session():
    try:
        s = stripe.checkout.Session.create(line_items=[{'price': STRIPE_PRICE_ID, 'quantity': 1}], mode='subscription', success_url=YOUR_DOMAIN + '/success?session_id={CHECKOUT_SESSION_ID}', cancel_url=YOUR_DOMAIN + '/')
        return redirect(s.url, code=303)
    except Exception as e: return str(e)

@app.route('/success')
@login_required
def success():
    user = User.query.get(current_user.id)
    user.is_premium = True
    db.session.commit()
    return render_template_string(f"<html lang='ar' dir='rtl'><head>{STYLE}</head><body><div class='container'><h1 style='color:#2ecc71;'>تم تفعيل الحساب المميز بنجاح 🎉</h1><p>يمكنك الآن البدء بحصر مشاريعك الهندسية بدون قيود.</p><a href='{url_for('index')}' class='btn' style='background:#2ecc71;'>الانتقال للوحة الحصر</a></div></body></html>")

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if not current_user.is_premium: return "خطأ بالصلاحية", 403
    file = request.files.get('file')
    if not file or file.filename == '': return "ملف غير صحيح", 400
    fname = file.filename
    ipath = os.path.join(UPLOAD_FOLDER, fname)
    file.save(ipath)
    oname = "BOQ_" + fname.rsplit('.', 1)[0] + ".xlsx"
    opath = os.path.join(UPLOAD_FOLDER, oname)
    
    success = run_cad(ipath, opath) if fname.lower().endswith('.dxf') else run_pdf(ipath, opath)
    if success: return send_file(opath, as_attachment=True, download_name=oname)
    return "حدث خطأ في المعالجة.", 500

def run_cad(ipath, opath):
    try:
        doc = ezdxf.readfile(ipath); msp = doc.modelspace(); recs = []
        for e in msp:
            lay = e.dxf.layer
            flr = lay if ("floor" in lay.lower() or "طابق" in lay) else "غير محدد"
            if e.dxftype() in ['LINE', 'LWPOLYLINE', 'POLYLINE']:
                lng = e.dxf.start.distance(e.dxf.end) if e.dxftype() == 'LINE' else e.length()
                recs.append({"الموقع": flr, "الطبقة": lay, "الوحدة": "متر طولي", "العنصر": e.dxftype(), "الكمية": round(lng, 2)})
            elif e.dxftype() == 'INSERT':
                recs.append({"الموقع": flr, "الطبقة": lay, "الوحدة": "عدد", "العنصر": e.dxf.name, "الكمية": 1})
        return save_xl(recs, opath)
    except: return False

def run_pdf(ipath, opath):
    try:
        doc = fitz.open(ipath); recs = []
        for p_num, p in enumerate(doc):
            lbl = f"صفحة {p_num + 1}"
            for path in p.get_drawings():
                for item in path["items"]:
                    if item == "l": recs.append({"الموقع": lbl, "الطبقة": "جرافيك متجه", "الوحدة": "متر طولي", "العنصر": "خطي هندسي", "الكمية": 0.5})
            for w in p.get_text("words"):
                txt = w[4].strip()
                if len(txt) > 1 and not txt.replace('.', '', 1).isdigit():
                    recs.append({"الموقع": lbl, "الطبقة": "رموز ونصوص", "الوحدة": "عدد", "العنصر": txt, "الكمية": 1})
        return save_xl(recs, opath)
    except: return False

def save_xl(recs, opath):
    if not recs: return False
    try:
        df = pd.DataFrame(recs).groupby(["الموقع", "الطبقة", "الوحدة", "العنصر"]).sum().reset_index()
        with pd.ExcelWriter(opath, engine='openpyxl') as w:
            df.to_excel(w, sheet_name='الكميات الإجمالية', index=False)
            ws = w.sheets['الكميات الإجمالية']
            ws.sheet_view.rightToLeft = True
            h_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            h_font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
            c_font = Font(name="Segoe UI", size=11)
            align = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.fill = h_fill; cell.font = h_font; cell.alignment = align
            for row in range(2, ws.max_row + 1):
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col); cell.font = c_font; cell.alignment = align; cell.border = border
